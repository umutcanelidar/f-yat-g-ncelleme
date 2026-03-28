import os
import uuid
import csv
import json
from io import BytesIO
import time
import queue
from flask import Flask, render_template, request, jsonify, send_file, Response
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from automation import IdeasoftBot

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(os.path.dirname(__file__), 'outputs')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Settings handling
SETTINGS_FILE = 'settings.json'
FIXED_SHOP_URL = "https://vhex10.myideasoft.com"

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f:
                return json.load(f)
        except:
            pass
    return {"username": "", "password": ""}

def save_settings(data):
    # Only save what's necessary (no shop_url)
    clean_data = {
        "username": data.get("username", ""),
        "password": data.get("password", "")
    }
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(clean_data, f)

# Global bot instance
bot = None

# Logging queue for UI
log_queue = queue.Queue(maxsize=100)

def push_log(message, level="info"):
    try:
        if log_queue.full():
            log_queue.get_nowait()
        log_queue.put_nowait({"message": message, "level": level})
    except:
        pass

@app.route('/api/logs')
def stream_logs():
    def generate():
        while True:
            try:
                log = log_queue.get(timeout=10)
                yield f"data: {json.dumps(log)}\n\n"
            except queue.Empty:
                yield "data: {\"message\": \"heartbeat\", \"level\": \"debug\"}\n\n"
            except Exception as e:
                break
    return Response(generate(), mimetype='text/event-stream')

# SKU sütunu anahtar kelimeleri
SKU_KEYWORDS = ['stok kodu', 'stok kod', 'sku', 'ürün kodu', 'urun kodu', 'kod',
                'barkod', 'barcode', 'code', 'product code', 'stock code']

# Fiyat sütunu anahtar kelimeleri (IdeaSoft - KDV'siz öncelikli)
PRICE_KEYWORDS_IDEASOFT = ['fiyat 1', 'birim fiyat', 'satış fiyatı', 'satis fiyati',
                           'fiyat', 'price', 'tutar']

# Fiyat sütunu anahtar kelimeleri (Yaşar Teknik)
PRICE_KEYWORDS_YASAR = ['fiyat', 'price', 'tutar', 'liste fiyat', 'birim fiyat',
                        'satış fiyatı', 'satis fiyati']


def find_column(columns, keywords):
    """Sütun listesinde anahtar kelimelere göre sütun bul."""
    cols_lower = {c: c.lower().strip().replace('*', '').replace('"', '') for c in columns}
    # Tam eşleşme öncelikli
    for kw in keywords:
        for orig, low in cols_lower.items():
            if low == kw:
                return orig
    # Kısmi eşleşme
    for kw in keywords:
        for orig, low in cols_lower.items():
            if kw in low:
                return orig
    return None


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/compare', methods=['POST'])
def compare():
    """İki dosyayı SKU'ya göre karşılaştırır."""
    if 'ideasoft_file' not in request.files or 'yasar_file' not in request.files:
        return jsonify({'error': 'Her iki dosya da gerekli'}), 400

    ideasoft_file = request.files['ideasoft_file']
    yasar_file = request.files['yasar_file']

    # İskonto oranı (varsayılan %45)
    iskonto = float(request.form.get('iskonto', 45))

    try:
        # Dosyaları oku
        df_idea = read_file(ideasoft_file)
        df_yasar = read_file(yasar_file)

        # SKU sütunlarını otomatik bul
        sku_col_idea = find_column(df_idea.columns, SKU_KEYWORDS)
        sku_col_yasar = find_column(df_yasar.columns, SKU_KEYWORDS)

        # SKU bulunamazsa ilk sütunu kullan (IdeaSoft bazen rastgele sütun adı veriyor)
        if not sku_col_idea:
            sku_col_idea = df_idea.columns[0]
        if not sku_col_yasar:
            sku_col_yasar = df_yasar.columns[0]

        # Fiyat sütunlarını otomatik bul
        price_col_idea = find_column(df_idea.columns, PRICE_KEYWORDS_IDEASOFT)
        price_col_yasar = find_column(df_yasar.columns, PRICE_KEYWORDS_YASAR)

        if not price_col_idea:
            return jsonify({'error': f'IdeaSoft dosyasında Fiyat sütunu bulunamadı. Mevcut sütunlar: {", ".join(df_idea.columns[:10])}'}), 400
        if not price_col_yasar:
            return jsonify({'error': f'Yaşar Teknik dosyasında Fiyat sütunu bulunamadı. Mevcut sütunlar: {", ".join(df_yasar.columns[:10])}'}), 400

        # Fiyat sütunlarını sayısala çevir
        df_idea[price_col_idea] = pd.to_numeric(
            df_idea[price_col_idea].astype(str).str.replace(',', '.').str.replace(r'[^\d.]', '', regex=True),
            errors='coerce'
        ).fillna(0)

        df_yasar[price_col_yasar] = pd.to_numeric(
            df_yasar[price_col_yasar].astype(str).str.replace(',', '.').str.replace(r'[^\d.]', '', regex=True),
            errors='coerce'
        ).fillna(0)

        # SKU sütunlarını normalize et
        df_idea['_sku_key'] = df_idea[sku_col_idea].astype(str).str.strip().str.lower()
        df_yasar['_sku_key'] = df_yasar[sku_col_yasar].astype(str).str.strip().str.lower()

        # Yaşar Teknik fiyatlarına iskonto uygula (liste fiyat → bayi fiyat)
        iskonto_carpan = 1 - (iskonto / 100)
        df_yasar['_bayi_fiyat'] = df_yasar[price_col_yasar] * iskonto_carpan

        # Yaşar Teknik fiyat haritası (iskontolu bayi fiyatı)
        yasar_price_map = dict(zip(df_yasar['_sku_key'], df_yasar['_bayi_fiyat']))
        # Liste fiyat haritası da tut (göstermek için)
        yasar_liste_map = dict(zip(df_yasar['_sku_key'], df_yasar[price_col_yasar]))

        # Ürün adı sütununu bul (varsa)
        name_keywords = ['ürün adı', 'urun adi', 'ürün', 'ad', 'name', 'product', 'açıklama', 'aciklama']
        name_col_idea = find_column(df_idea.columns, name_keywords)

        # Eşleştirme ve fark tespiti
        changes = []
        matched_count = 0
        changed_count = 0
        matched_yasar_keys = set()

        for idx, row in df_idea.iterrows():
            key = row['_sku_key']
            if key and key != 'nan' and key in yasar_price_map:
                matched_count += 1
                matched_yasar_keys.add(key)
                old_price = row[price_col_idea]
                new_price = yasar_price_map[key]
                if abs(old_price - new_price) > 0.01:
                    changed_count += 1
                    product_label = str(row[name_col_idea]) if name_col_idea else str(row[sku_col_idea])
                    changes.append({
                        'idx': idx,
                        'sku': str(row[sku_col_idea]),
                        'product_name': product_label,
                        'old_price': round(float(old_price), 2),
                        'liste_fiyat': round(float(yasar_liste_map.get(key, 0)), 2),
                        'new_price': round(float(new_price), 2)
                    })

        not_matched_ideasoft = len(df_idea) - matched_count
        not_matched_yasar = len(df_yasar) - len(matched_yasar_keys)

        # Eşleşmeyen IdeaSoft ürünleri (SKU'su Yaşar'da bulunamayanlar)
        unmatched_idea_list = []
        for idx, row in df_idea.iterrows():
            key = row['_sku_key']
            if not key or key == 'nan' or key not in yasar_price_map:
                product_label = str(row[name_col_idea]) if name_col_idea else ''
                unmatched_idea_list.append({
                    'sku': str(row[sku_col_idea]),
                    'product_name': product_label,
                    'price': round(float(row[price_col_idea]), 2) if row[price_col_idea] else 0
                })

        # Eşleşmeyen Yaşar Teknik ürünleri
        unmatched_yasar_list = []
        for idx, row in df_yasar.iterrows():
            key = row['_sku_key']
            if key not in matched_yasar_keys:
                unmatched_yasar_list.append({
                    'sku': str(row[sku_col_yasar]),
                    'price': round(float(row[price_col_yasar]), 2) if row[price_col_yasar] else 0
                })

        # Sonuç Excel'ini oluştur
        output_filename = f'sonuc_{uuid.uuid4().hex[:8]}.xlsx'
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

        df_output = df_idea.drop(columns=['_sku_key'])
        df_output.to_excel(output_path, index=False, engine='openpyxl')

        # Openpyxl ile formatlama
        wb = load_workbook(output_path)
        ws = wb.active

        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        red_font = Font(color='CC0000', bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        price_col_idx = list(df_output.columns).index(price_col_idea) + 1

        new_col_yasar = ws.max_column + 1
        new_col_status = ws.max_column + 2

        ws.cell(row=1, column=new_col_yasar, value='Yaşar Teknik Fiyat').fill = header_fill
        ws.cell(row=1, column=new_col_yasar).font = header_font
        ws.cell(row=1, column=new_col_yasar).alignment = Alignment(horizontal='center')

        ws.cell(row=1, column=new_col_status, value='Durum').fill = header_fill
        ws.cell(row=1, column=new_col_status).font = header_font
        ws.cell(row=1, column=new_col_status).alignment = Alignment(horizontal='center')

        changed_indices = {c['idx'] for c in changes}
        for change in changes:
            row_num = change['idx'] + 2
            price_cell = ws.cell(row=row_num, column=price_col_idx)
            price_cell.fill = red_fill
            price_cell.font = red_font

            yasar_cell = ws.cell(row=row_num, column=new_col_yasar, value=change['new_price'])
            yasar_cell.font = Font(bold=True)

            status_cell = ws.cell(row=row_num, column=new_col_status, value='FİYAT DEĞİŞTİ')
            status_cell.fill = red_fill
            status_cell.font = red_font

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        wb.save(output_path)

        return jsonify({
            'total_ideasoft': len(df_idea),
            'total_yasar': len(df_yasar),
            'total_matched': matched_count,
            'price_changed': changed_count,
            'not_matched_ideasoft': not_matched_ideasoft,
            'not_matched_yasar': not_matched_yasar,
            'iskonto': iskonto,
            'sku_col_idea': sku_col_idea,
            'sku_col_yasar': sku_col_yasar,
            'price_col_idea': price_col_idea,
            'price_col_yasar': price_col_yasar,
            'changes': changes[:200],
            'unmatched_ideasoft': unmatched_idea_list[:200],
            'unmatched_yasar': unmatched_yasar_list[:200],
            'filename': output_filename
        })

    except Exception as e:
        return jsonify({'error': f'Karşılaştırma hatası: {str(e)}'}), 500


# ========== AUTOMATION ROUTES ==========

@app.route('/api/settings', methods=['GET', 'POST'])
def api_settings():
    if request.method == 'POST':
        data = request.json
        save_settings(data)
        return jsonify({"status": "success"})
    return jsonify(load_settings())

@app.route('/api/login', methods=['POST'])
def api_login():
    global bot
    settings = load_settings()
    
    if bot: bot.close()
    bot = IdeasoftBot(shop_url=FIXED_SHOP_URL)
    
    # We pass credentials to the login method
    success = bot.login(username=settings.get("username"), password=settings.get("password"))
    if success:
        return jsonify({"status": "success", "message": "Giriş başarılı ve oturum kaydedildi."})
    return jsonify({"error": "Giriş başarısız oldu veya tarayıcı kapatıldı."}), 401

@app.route('/api/update-price', methods=['POST'])
def api_update_price():
    global bot
    data = request.json
    sku = data.get('sku')
    price = data.get('price')
    
    if not bot:
        bot = IdeasoftBot(shop_url=FIXED_SHOP_URL, log_callback=push_log)
    
    success, message = bot.update_price(sku, price)
    if success:
        return jsonify({"status": "success"})
    return jsonify({"error": message}), 200 # Return 200 so JS can handle the error message properly

@app.route('/api/bulk-update', methods=['POST'])
def api_bulk_update():
    global bot
    # This would be a long-running process, ideally async.
    # For now, we'll do it sequentially and return progress in a later version or via logs.
    data = request.json
    products = data.get('products', []) # List of {sku, price}
    
    if not bot:
        bot = IdeasoftBot(shop_url=FIXED_SHOP_URL)
    
    results = []
    for p in products:
        success, message = bot.update_price(p['sku'], p['price'])
        results.append({"sku": p['sku'], "success": success, "message": message})
        
    return jsonify({"status": "completed", "results": results})

@app.route('/download/<filename>')
def download(filename):
    """Sonuç dosyasını indir."""
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'Dosya bulunamadı'}), 404
    return send_file(filepath, as_attachment=True, download_name='fiyat_karsilastirma_sonuc.xlsx')


def read_file(file_obj):
    """Excel veya CSV dosyasını oku."""
    filename = file_obj.filename
    ext = filename.rsplit('.', 1)[-1].lower()
    if ext == 'csv':
        raw = file_obj.read()
        for encoding in ['utf-8', 'latin-1', 'cp1254', 'iso-8859-9']:
            for sep in [';', ',', '\t', '|']:
                try:
                    df = pd.read_csv(BytesIO(raw), sep=sep, encoding=encoding,
                                     quoting=csv.QUOTE_NONE, on_bad_lines='skip', comment='#')
                    if len(df.columns) > 1:
                        return df
                except:
                    continue
        return pd.read_csv(BytesIO(raw), sep=None, engine='python',
                           encoding='latin-1', quoting=csv.QUOTE_NONE, on_bad_lines='skip', comment='#')
    else:
        return pd.read_excel(file_obj)


if __name__ == '__main__':
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    print('\n🚀 Fiyat Karşılaştırma Paneli başlatıldı!')
    print(f'📍 Bu bilgisayar: http://localhost:5050')
    print(f'🌐 Ağdaki cihazlar: http://{local_ip}:5050\n')
    app.run(debug=True, port=5050, host='0.0.0.0')
